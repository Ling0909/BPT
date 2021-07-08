from Common.base_page import BasePage
from PageLocators.locators_portfolioPlanPage import LocatorsPortfolioPlanPage as lppp
from selenium.webdriver.common.action_chains import ActionChains
from Common.datas_stor import DatasStor
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from Common.path_object import file_create_portfoliio
import time

class PagePortfolioPlan(BasePage):

    def find_by_portfolioNo(self,CostFundingName,portfolioNo,lineCode):
        """
        :param CostFundingName: Country Portfolio & Plan页面Cost & Funding模块下的CostFunding名字，
                            在程序中经常用CostFunding_Value代表CostFunding的平均值
        :param portfolioNo:
        :param lineCode:
        :return:示例 CostFunding_ValueDict={'GEO Funding(Others) 1 ($)': '1.00', 'JUL.': '1.00', 'AUG.': '1.00', 'SEP.': '1.00'}
        """
        doc='portfolio页面操作'
        self.wait_eleVisible(lppp.ele_iframe,50,doc=doc)
        self.switch_iframe(self.get_element(lppp.ele_iframe,doc))
        self.wait_eleVisible(lppp.input_portfolioNo,50,doc=doc)
        self.input_text(lppp.input_portfolioNo,portfolioNo,doc) #在Portfolio No.输入框输入条件
        self.click_element(lppp.button_search,doc) #点击搜索按钮
        time.sleep(1)
        #获取列表中国家名字，并按照国家名字获取到GEO和REGION
        baseInfo={}
        self.wait_eleVisible(lppp.list_results_portfolio,10,doc=doc) # 等待按照portfolio No.查询结束
        Plan_Cycle=self.get_text(lppp.planCycle_on_portfolioList,doc)
        setattr(DatasStor,'Plan_Cycle',Plan_Cycle)
        time.sleep(1)
        countryName=self.get_text(lppp.countryValue_on_portfolioList,doc)  # 获取国家名称
        self.click_element(lppp.input_country)
        self.input_text(lppp.input_country_in_list,countryName,doc)
        time.sleep(1)
        GEO=self.get_text(lppp.ele_geo_in_list,doc)
        RegionName=self.get_text(lppp.ele_region_in_list,doc)

        if self.get_element(lppp.ele_subRegion_in_list,doc):
            baseInfo['Sub_Region']=self.get_text(lppp.ele_subRegion_in_list,doc)
            baseInfo['Region']=RegionName
        else:
            baseInfo['Region']=RegionName
        baseInfo['GEO']=GEO
        time.sleep(1)

        #继续选择输入筛选条件，筛选portfolio信息
        self.click_element(lppp.ele_portfolio,doc)# 点击搜索结果中的Portfolio No.值
        self.wait_eleVisible(lppp.ele_lineCode_OK,doc=doc)
        self.input_text(lppp.input_product_lineCode,lineCode,doc)
        time.sleep(3)
        eles_lineCode=self.get_elements(lppp.eles_lineCode,doc)
        #鼠标操作
        for txt_lineCode in eles_lineCode:
            if txt_lineCode.text==lineCode:
                self.double_click_element(txt_lineCode,doc)
                break

        #获取baseinfo{}其他信息
        self.wait_eleVisible(lppp.ele_load_img,wait_times=100,doc=doc)  # 等待在portfolio界面出现遮罩层
        self.wait_eleVisible(lppp.ele_country, wait_times=100,doc=doc)  # 等待国家字段元素出现，证明页面加载完成
        baseInfo['COUNTRY']=self.get_element_attribute(lppp.ele_country,'value')
        if baseInfo['GEO']=='AP':#只有AP区域的国家会在界面显示AP_RTM字段，其他区域不显示在界面并默认为TS
            baseInfo['AP_RTM']=self.get_element_attribute(lppp.ele_AP_RTM,'value')
        else:
            baseInfo['AP_RTM'] ='TS'

        baseInfo['PRODUCT_GROUP']=self.get_element_attribute(lppp.ele_productGroup,'value',doc)
        baseInfo['PRODUCT_CATEGORY'] =self.get_element_attribute(lppp.ele_productCategory,'value',doc)
        baseInfo['PRODUCT_SERIES']=self.get_element_attribute(lppp.ele_productSeries,'value',doc)
        baseInfo['PRODUCT_FAMILY']=self.get_element_attribute(lppp.ele_productFamily,'value',doc)
        baseInfo['PN'] = self.get_element_attribute(lppp.ele_PN, 'value', doc)
        while True:
            self.driver.execute_script('window.scrollBy(0,100)')
            time.sleep(2)
            self.click_element(lppp.button_packUp,doc)#点击Description的下拉箭头
            break
            # 获取componentV信息
        while True:
            # cv=self.driver.find_element_by_id('collapseCV')
            self.driver.execute_script('window.scrollBy(0,100)')
            components=self.driver.find_elements_by_css_selector('#collapseCV .col-xs-6 div[class*="col-xs-9"] input[value]')
            componentsV=[]
            if components:
                for one in components:
                    con=one.get_attribute('value')
                    if con in componentsV:
                        continue
                    componentsV.append(con)
                break
            break
        baseInfo['Component V']=componentsV
        # 关闭组件信息
        self.wait_eleVisible(lppp.button_packUp,doc=doc)
        self.click_element(lppp.button_packUp,doc)
        #获取baseInfo['Channel Price($)']的值
        if baseInfo['COUNTRY']=='Brazil':
            baseInfo['Channel Price($)']='0'
        else:
            ele=self.get_element(lppp.ele_ChannelPrice,doc)
            self.driver.execute_script('arguments[0].scrollIntoView(false);', ele)  #移动到channelPrice的底部
            baseInfo['Channel Price($)']=ele.get_attribute('value')
        self.log.logger_info('最后得到baseInfo的值是：{}'.format(baseInfo))
        setattr(DatasStor, 'baseInfo', baseInfo)  # 将baseInfo赋值给datas_stor下
        # 点击Calculate按钮
        self.click_element(lppp.button_Calculate,doc)
        # 等待元素出现
        self.wait_eleVisible(lppp.ele_load_img, 300, doc=doc)
        self.wait_elePresence(lppp.ele_load_img_disappear, 100, doc=doc)

        #获取“CostFundingName”相关的值
        CostFunding_ValueDict={}
        ele_CostFundingName=None#FundingName元素定位表达式
        eles_CostFunding_mouths =None #月份元素定位表达式
        eles_CostFunding_mouthValues =None #月份值元素定位表达式

        #通过不同的值选择不同的CostFunding元素，后期增加CostFunding元素的时候，需要在这里与locators_portfolioPlanPage页面添加相应的元素数据
        if CostFundingName=='GEOFundingOthers1':
            if baseInfo['COUNTRY']=='Brazil':
                ele_CostFundingName=lppp.ele_FundingOthers1_Brazil
                eles_CostFunding_mouths=lppp.eles_FundingOthers1_Brazil_months
                eles_CostFunding_mouthValues=lppp.eles_FundingOthers1_Brazil_mouthsValues
            else:
                ele_CostFundingName =lppp.ele_FundingOthers1
                eles_CostFunding_mouths=lppp.eles_FundingOthers1_months
                eles_CostFunding_mouthValues=lppp.eles_FundingOthers1_mouthsValues
        elif CostFundingName=='GEOFundingOthers2':
            ele_CostFundingName=lppp.ele_FundingOthers2
            eles_CostFunding_mouths=lppp.eles_FundingOthers2_months
            eles_CostFunding_mouthValues=lppp.eles_FundingOthers2_mouthsValues
        elif CostFundingName=='GEOFundingOthers3':
            ele_CostFundingName=lppp.ele_FundingOthers3
            eles_CostFunding_mouths=lppp.eles_FundingOthers3_months
            eles_CostFunding_mouthValues=lppp.eles_FundingOthers3_mouthsValues
        elif CostFundingName=='GEOFundingOthers4':
            ele_CostFundingName=lppp.ele_FundingOthers4
            eles_CostFunding_mouths=lppp.eles_FundingOthers4_months
            eles_CostFunding_mouthValues=lppp.eles_FundingOthers4_mouthsValues
        elif CostFundingName=='Funding1_CPU':
            ele_CostFundingName=lppp.ele_Funding1_CPU
            eles_CostFunding_mouths=lppp.eles_Funding1_CPU_months
            eles_CostFunding_mouthValues=lppp.eles_Funding1_CPU_mouthsValues
        elif CostFundingName=='Funding2_HDD_SSHD_SSD':
            ele_CostFundingName=lppp.ele_Funding2_HDD_SSHD_SSD
            eles_CostFunding_mouths=lppp.eles_Funding2_HDD_SSHD_SSD_months
            eles_CostFunding_mouthValues=lppp.eles_Funding2_HDD_SSHD_SSD_mouthsValues
        elif CostFundingName=='Funding3_Others':
            ele_CostFundingName=lppp.ele_Funding3_Others
            eles_CostFunding_mouths=lppp.eles_Funding3_Others_months
            eles_CostFunding_mouthValues=lppp.eles_Funding3_Others_mouthsValues
        elif CostFundingName=='APCostCredit':
            ele_CostFundingName=lppp.ele_APCostCredit
            eles_CostFunding_mouths=lppp.eles_APCostCredit_months
            eles_CostFunding_mouthValues=lppp.eles_APCostCredit_mouthsValues
        elif CostFundingName=='CountryAdjustment':
            if baseInfo['COUNTRY']=='Brazil':
                ele_CostFundingName=lppp.ele_CountryAdjustment_Brazil
                eles_CostFunding_mouths=lppp.eles_CountryAdjustment_Brazil_months
                eles_CostFunding_mouthValues=lppp.eles_CountryAdjustment_Brazil_mouthsValues
            else:
                ele_CostFundingName =lppp.ele_CountryAdjustment
                eles_CostFunding_mouths=lppp.eles_CountryAdjustment_months
                eles_CostFunding_mouthValues=lppp.eles_CountryAdjustment_mouthsValues
        elif CostFundingName == 'NonBMCUplift':
            ele_CostFundingName = lppp.ele_NonBMCUplift
            eles_CostFunding_mouths = lppp.eles_NonBMCUplift_months
            eles_CostFunding_mouthValues = lppp.eles_NonBMCUplift_mouthsValues
        elif CostFundingName == 'PNAssessment':
            ele_CostFundingName = lppp.ele_PNAssessment
            eles_CostFunding_mouths = lppp.eles_PNAssessment_months
            eles_CostFunding_mouthValues = lppp.eles_PNAssessment_mouthsValues
        elif CostFundingName == 'CountryAdjustment2':
            ele_CostFundingName = lppp.ele_CountryAdjustment2
            eles_CostFunding_mouths = lppp.eles_CountryAdjustment2_months
            eles_CostFunding_mouthValues = lppp.eles_CountryAdjustment2_mouthsValues


        CostFunding_Value=self.get_element_attribute(ele_CostFundingName,'value',doc)

        # 获取三个月份的值（月份值和数字值）
        CostFunding_ValueDict['CostFunding_value']='%.2f' %float(CostFunding_Value)
        months=self.get_elements(eles_CostFunding_mouths)
        monthsValue=self.get_elements(eles_CostFunding_mouthValues)

        CostFunding_ValueDict[months[0].text]='%.2f' %float((monthsValue[0].get_attribute('value')))
        CostFunding_ValueDict[months[1].text]='%.2f' %float((monthsValue[1].get_attribute('value')))
        CostFunding_ValueDict[months[2].text]='%.2f' %float((monthsValue[2].get_attribute('value')))


        return CostFunding_ValueDict

    def get_datas_excel(self):
        datas_portfolio=[]
        filename=file_create_portfoliio
        book=load_workbook(filename)
        # 获取第一张sheet表单
        sh =book['datas']
        for nrow in range(2,sh.max_row+1):
            create_portfolioDict={}
            create_portfolioDict['CostFundingDataName']=sh.cell(nrow,1).value
            create_portfolioDict['Country']=sh.cell(nrow,2).value
            create_portfolioDict['Product_Group']=sh.cell(nrow,3).value
            create_portfolioDict['Plan_Cycle'] =sh.cell(nrow,4).value
            create_portfolioDict['Portfolio_Name']=sh.cell(nrow,5).value
            datas_portfolio.append(create_portfolioDict)
            portfolioName_mark=str(int(sh.cell(sh.max_row,5).value[-3:])+nrow-1).zfill(3)
            portfolioName=f'auto_create_test{portfolioName_mark}'
            self.write_excel(filename,'datas',nrow,5,portfolioName)
        return datas_portfolio

    def add_portfolio_datas(self,data):
        doc='增加portfolio数据'
        #进入第一层的iframe
        self.wait_eleVisible(lppp.ele_iframe,doc=doc)
        self.driver.switch_to.frame(self.driver.find_element(*lppp.ele_iframe))
        #开始添加数据

        self.wait_eleVisible(lppp.input_portfolioname, doc=doc)
        self.input_text(lppp.input_portfolioname, data['Portfolio_Name'])
        self.click_element(lppp.button_search)
        time.sleep(3)
        if self.is_eleExist(lppp.button_del,doc=doc):#判断是否存在
            self.click_element(lppp.ele_portfolio,doc=doc)
            time.sleep(1)
            if self.is_eleExist(lppp.ele_product_title_result,doc=doc):
                ele_lineCode=self.driver.find_element(*lppp.ele_product_title_result)
                # 鼠标操作
                ac=ActionChains(self.driver)
                ac.double_click(ele_lineCode).perform()
                self.wait_eleVisible(lppp.ele_lineCode_baseInfo)
                lineCode=self.get_element_attribute(lppp.ele_lineCode_baseInfo,'value',doc)
            else:
                # 通过MTM界面为portfolio维护lineCode
                self.wait_eleVisible(lppp.button_add_lineCode, doc=doc)
                self.click_element(lppp.button_add_lineCode, doc=doc)
                self.click_element(lppp.button_add_from_MTM, doc=doc)
                self.wait_eleVisible(lppp.ele_check_box, doc=doc)
                self.click_element(lppp.ele_check_box, doc=doc)
                self.click_element(lppp.button_sava, doc=doc)
                self.wait_eleVisible(lppp.ele_product_title_result, doc=doc)  # 等待界面的元素加载完毕

                ele_lineCode=self.get_element(lppp.ele_product_title_result,doc)
                # 鼠标操作
                ac=ActionChains(self.driver)
                ac.double_click(ele_lineCode).perform()
                self.wait_eleVisible(lppp.ele_lineCode_baseInfo)
                lineCode=self.get_element_attribute(lppp.ele_lineCode_baseInfo,'value', doc)


        else:
            self.click_element(lppp.button_create,doc)#点击create按钮后，弹出维护数据的弹窗
            self.wait_eleVisible(lppp.ele_iframe_pop,doc=doc)
            self.driver.switch_to.frame(self.driver.find_element(*lppp.ele_iframe_pop))
            #在弹窗中维护portfolio的信息
            self.wait_eleVisible(lppp.input_country_pop,doc=doc)
            self.click_element(lppp.input_country_pop,doc)
            self.input_text(lppp.input_search_country_pop,data['Country'],doc)
            self.driver.find_element_by_xpath('//ul[@id="txtCountrytreeId"]//span[text()="{}"]'.format(data['Country'])).click()
            Select(self.driver.find_element(*lppp.select_product_group_pop)).select_by_visible_text(data['Product_Group'])
            Select(self.driver.find_element(*lppp.select_plan_cycle_pop)).select_by_index(data['Plan_Cycle'])
            self.input_text(lppp.input_portfolioName_pop,data['Portfolio_Name'])
            self.driver.find_element(*lppp.button_create_pop).click()
            time.sleep(2)


            self.driver.switch_to.parent_frame()


            # ele_portfolioNo=self.driver.find_element_by_xpath('//div[@id="borderBox"]//td[contains(text(),"{}")]/..//td[2]'.format(data['Portfolio_Name']))
            # portfolioNoDict['portfolioNo']=ele_portfolioNo.text#将新增加的portfolioNo值放到字典里面
            # ele_portfolioNo.click()
            #通过MTM界面为portfolio维护lineCode
            self.wait_eleVisible(lppp.button_add_lineCode,doc=doc)
            self.click_element(lppp.button_add_lineCode,doc=doc)
            self.click_element(lppp.button_add_from_MTM,doc=doc)
            self.wait_eleVisible(lppp.ele_check_box,doc=doc)
            self.click_element(lppp.ele_check_box, doc=doc)
            self.click_element(lppp.button_sava, doc=doc)
            self.wait_eleVisible(lppp.ele_product_title_result, doc=doc)  # 等待界面的元素加载完毕

            ele_lineCode=self.driver.find_element(*lppp.ele_product_title_result)
            # 鼠标操作
            ac=ActionChains(self.driver)
            ac.double_click(ele_lineCode).perform()
            self.wait_eleVisible(lppp.ele_lineCode_baseInfo)
            lineCode=self.get_element_attribute(lppp.ele_lineCode_baseInfo,'value', doc)
        portfolioNoDict={}
        portfolioNoDict['lineCode']=lineCode#将lineCode放到字典里
        portfolioNoDict['portfolioNo'] =lineCode[:-7]
        portfolioNoDict['CostFundingDataName']=data['CostFundingDataName']
        return portfolioNoDict
        # #分类整理添加的数据
        # if data['CostFundingDataName']=='datas_GEOFundingOthers1':
        #     datas_GEOFundingOthers1=[]
        #     datas_GEOFundingOthers1.append(portfolioNoDict)
        #     setattr(DatasStor,'datas_GEOFundingOthers1',datas_GEOFundingOthers1)
        # elif data['CostFundingDataName']=='datas_GEOFundingOthers1_GEO_LA':
        #     datas_GEOFundingOthers1_GEO_LA=[]
        #     datas_GEOFundingOthers1_GEO_LA.append(portfolioNoDict)
        #     setattr(DatasStor, 'datas_GEOFundingOthers1_GEO_LA',datas_GEOFundingOthers1_GEO_LA)
        # elif data['CostFundingDataName']=='datas_GEOFundingOthers2':
        #     datas_GEOFundingOthers2=[]
        #     datas_GEOFundingOthers2.append(portfolioNoDict)
        #     setattr(DatasStor, 'datas_GEOFundingOthers2',datas_GEOFundingOthers2)
        # elif data['CostFundingDataName']=='datas_GEOFundingOthers3':
        #     datas_GEOFundingOthers3=[]
        #     datas_GEOFundingOthers3.append(portfolioNoDict)
        #     setattr(DatasStor, 'datas_GEOFundingOthers3', datas_GEOFundingOthers3)
        # elif data['CostFundingDataName']=='datas_GEOFundingOthers4':
        #     datas_GEOFundingOthers4=[]
        #     datas_GEOFundingOthers4.append(portfolioNoDict)
        #     setattr(DatasStor, 'datas_GEOFundingOthers4', datas_GEOFundingOthers4)
        # elif data['CostFundingDataName']=='datas_Funding1_CPU':
        #     datas_Funding1_CPU=[]
        #     datas_Funding1_CPU.append(portfolioNoDict)
        #     setattr(DatasStor, 'datas_Funding1_CPU',datas_Funding1_CPU)
        # elif data['CostFundingDataName']=='datas_Funding2_HDD_SSHD_SSD':
        #     datas_Funding2_HDD_SSHD_SSD=[]
        #     datas_Funding2_HDD_SSHD_SSD.append(portfolioNoDict)
        #     setattr(DatasStor, 'datas_Funding2_HDD_SSHD_SSD',datas_Funding2_HDD_SSHD_SSD)
        # elif data['CostFundingDataName']=='datas_Funding3_Others':
        #     datas_Funding3_Others=[]
        #     datas_Funding3_Others.append(portfolioNoDict)
        #     setattr(DatasStor, 'datas_Funding3_Others', datas_Funding3_Others)
    def del_portfolio_datas(self,data):
        doc='删除portfolio数据'
        # 进入第一层的iframe
        self.wait_eleVisible(lppp.ele_iframe, doc=doc)
        self.driver.switch_to.frame(self.driver.find_element(*lppp.ele_iframe))
        # 开始添加数据

        self.wait_eleVisible(lppp.input_portfolioname,doc=doc)
        self.clear_element_text(lppp.input_portfolioname,doc)
        self.input_text(lppp.input_portfolioname, data['Portfolio_Name'])
        self.click_element(lppp.button_search)
        self.wait_eleVisible(lppp.button_del,doc=doc)
        self.click_element(lppp.button_del,doc)
        time.sleep(1)
        self.wait_eleVisible(lppp.button_del_ok,doc=doc)
        self.click_element(lppp.button_del_ok,doc)
        time.sleep(2)
        self.driver.switch_to.parent_frame()