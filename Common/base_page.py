# import time
from openpyxl import load_workbook
import win32com.client
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from Common.logger import Logger
from Common.path_object import *
import datetime
class BasePage:
    def __init__(self,driver):
        self.driver=driver
        self.log = Logger()

    def wait_eleVisible(self, locator, wait_times=30, poll_frequency=0.5, doc=""):
        """

        :param locator:元素定位。元祖形式。（元素定位类型、元素定位方式）
        :param times:等待时间
        :param poll_frequency:等待过程中的查询周期
        :param doc:传给截屏函数的名称，用户自定义，如果没有定义，就传给默认值空
        """
        #Logger().logger_info('等待元素{0}可见'.format(locator))
        try:
            #start_time = datetime.datetime.now()
            WebDriverWait(self.driver, wait_times).until(EC.visibility_of_element_located(locator))
            #end_time = datetime.datetime.now()
            #wait_time = (end_time - start_time).seconds
            # self.log.logger_info("{0}: 元素 {1} 已可见,等待起始时间：{2},等待结束时间：{3}.等待时长：{4}".format(doc, locator, start_time, end_time, wait_time))
        except:
            self.log.logger_error('等待元素{0}可见失败'.format(locator))
            self.save_screenshot(doc)
            raise

    def wait_elePresence(self, locator, wait_times=30, poll_frequency=0.5, doc=""):
        """

        :param locator:元素定位。元祖形式。（元素定位类型、元素定位方式）
        :param times:等待时间
        :param poll_frequency:等待过程中的查询周期
        :param doc:传给截屏函数的名称，用户自定义，如果没有定义，就传给默认值空
        """
        Logger().logger_info('等待元素{0}存在'.format(locator))
        try:
            start_time = datetime.datetime.now()
            WebDriverWait(self.driver, wait_times).until(EC.presence_of_element_located(locator))
            end_time = datetime.datetime.now()
            wait_time = (end_time - start_time).seconds
            # self.log.logger_info("{0}: 元素 {1} 已存在,等待起始时间：{2},等待结束时间：{3}.等待时长：{4}".format(doc, locator, start_time, end_time, wait_time))
        except:
            self.log.logger_error('等待元素{0}可见失败'.format(locator))
            self.save_screenshot(doc)
            raise


#查找单个元素
    def get_element(self,locator,doc=""):
        # self.log.logger_info('开始查找元素{0}'.format(locator))
        try:
            return self.driver.find_element(*locator)
        except:
            self.log.logger_error('查找元素失败！')
            self.save_screenshot(doc)
            raise

    #查找多个元素
    def get_elements(self, locator, doc=""):
        # self.log.logger_info('开始批量查找元素{0}'.format(locator))
        try:
            return self.driver.find_elements(*locator)
        except:
            self.log.logger_error('批量查找元素失败！')
            self.save_screenshot(doc)
            raise
    #点击操作
    def click_element(self,locator,doc=""):
        #找元素
        # ele=self.get_element(locator,doc=doc)
        # self.log.logger_info("{0} 点击元素：{1}".format(doc, locator))
        #元素操作
        try:
            self.driver.find_element(*locator).click()
        except:
            self.log.logger_error('点击元素失败！')
            self.save_screenshot(doc)
            raise

    #双击操作
    def double_click_element(self,double_click_reference,doc=''):
        """

        :param double_click_reference: 比如：=driver.find_element_by_xpath("xxxx")
        :param doc:
        :return:
        """
        ac = ActionChains(self.driver)
        self.log.logger_info(f'在页面{doc}中双击元素')
        try:
            ac.double_click(double_click_reference).perform()
        except:
            self.log.logger_error("双击元素操作失败！！！")
            # 截图
            self.save_screenshot(doc)
            raise
    #输入操作
    def input_text(self,locator,text,doc=""):
        ele=self.get_element(locator,doc=doc)
        # self.log.logger_info('{0}：元素：{1}输入内容{2}'.format(doc,locator,text))
        try:
            ele.send_keys(text)
        except:
            self.log.logger_error('元素输入操作失败！')
            raise

    #获取元素的文本内容
    def get_text(self,locator,doc=""):
        ele = self.get_element(locator, doc=doc)
        # self.log.logger_info('{0}：获取元素{1}的内容'.format(doc, locator))
        try:
            return ele.text
        except:
            self.log.logger_error('获取元素内容失败！')
            raise
    #清空输入框的内容
    def clear_element_text(self,locator,doc=""):
        ele=self.get_element(locator,doc=doc)
        # self.log.logger_info('{0}：情况输入框的内容'.format(doc))
        try:
            ele.clear()
        except:
            self.log.logger_error('清空元素内容失败！')
            raise
    #获取元素的属性
    def get_element_attribute(self,locator,attr,doc=""):
        ele = self.get_element(locator, doc)
        # self.log.logger_info("{0}: 获取元素：{1} 的属性：{2}".format(doc, locator, attr))
        try:
            ele_attr = ele.get_attribute(attr)
            # self.log.logger_info("元素：{0} 的属性 {1}  值为：{2}".format(locator, attr, ele_attr))
            return ele_attr
        except:
            self.log.logger_error("获取元素的属性失败！！！")
            # 截图
            self.save_screenshot(doc)
            raise
    #找元素
    # 元素存在则为True，否则为False
    def is_eleExist(self, locator, timeout=10, doc=""):
        # self.log.logger_info("在页面 {0} 中是否存在元素：{1}".format(doc, locator))
        try:
            WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located(locator))
            # self.log.logger_info(" {0} 秒内页面 {1} 中存在元素: {2}".format(timeout, doc, locator))
            return True
        except:
            self.log.logger_error(" {0} 秒内页面 {1} 中不存在元素: {2}".format(timeout, doc, locator))
            return False

    #alert处理
    def alert_action(self,action="accept"):
        pass

    #iframe切换
    def switch_iframe(self,ifram_reference,doc=''):
        """

        :param ifram_reference: 三种形式：name属性、iframe下角标、
            driver.switch_to.frame('frame_name')
            driver.switch_to.frame(1)
            driver.switch_to.frame(driver.find_elements_by_tag_name("iframe")[0])
        :param doc:
        :return:
        """
        self.log.logger_info(f'在页面{doc}中切换到iframe{ifram_reference}')

        try:
            self.driver.switch_to.frame(ifram_reference)
        except:
            self.log.logger_error("切换iframe窗口操作失败！！！")
            # 截图
            self.save_screenshot(doc)
            raise

    #滚动到对应页面底部处理
    def scroll_element_to_bottom(self,locator,doc=''):
        # self.log.logger_info("在页面 {0} 中滚动元素：{1}到页面底部".format(doc, locator))
        ele = self.get_element(locator, doc)
        try:
            self.driver.execute_script("arguments[0].scrollIntoView(false);",ele)
        except:
            self.log.logger_error("元素滚动到底部操作失败！！！")
            # 截图
            self.save_screenshot(doc)
            raise

    #select/option下拉拉列表选择
    def select_selectBox(self,locator,OptionValue,doc=''):
        """
        :param locator: 定位select下拉框位置
        :param OptionValue: 要选择的option元素的值
        :return: 无
        """
        ele_select= self.get_element(locator,doc)

        try:
            Select(ele_select).select_by_value(OptionValue)
        except:
            self.log.logger_error("选择select/option下拉框失败")
            # 截图
            self.save_screenshot(doc)
            raise

    #窗口切换
    #上传文件
    def up_load_file(self,importfile):
        shell = win32com.client.Dispatch("WScript.Shell")
        shell.Sendkeys(f'd:\\import\\{importfile}' + '\r\n')
        time.sleep(1)
    #获取最新下载的文件
    def get_latest_fileName(self):
        test_report=r"C:\Users\Administrator\Downloads"
        lists = os.listdir(test_report)  # 列出目录的下所有文件和文件夹保存到lists
        lists.sort(key=lambda fn: os.path.getmtime(test_report + "\\" + fn))  # 按时间排序
        file_new = os.path.join(test_report, lists[-1])  # 获取最新的文件保存到file_new
        return file_new

    def save_screenshot(self, doc):
        file_path = screen_path + "/{0}_{1}.png".format(doc, time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime()))
        try:
            self.driver.save_screenshot(file_path)
            # self.log.logger_info('截图成功！')
        except AssertionError as e:
            self.log.logger_error('截图失败！！')
            raise e
    #保存数据到Excel对应的单元格中
    def write_excel(self,file_name, sheet_name, i, j, data):  # 传入文件存储路径、excel的sheet名称、以及要插入的数据
        """
        :param file_name: 文件的绝对路径
        :param sheet_name: sheet名
        :param i: 行
        :param j: 列
        :param data:要写入单元格的数据
        :return:
        """
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, j).value = data
        wb.save(file_name)

